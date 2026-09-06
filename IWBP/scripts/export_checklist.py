"""Export docs/气象服务席检查单工作项.xlsx -> docs/checklist-data.json + embed in index.html.

修订检查单：改 xlsx 后在本机执行 ``python scripts/export_checklist.py`` 即可（会同步更新 index.html 内嵌 JSON）。
"""
import datetime as dt
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
path = ROOT / "docs" / "气象服务席检查单工作项.xlsx"
HTML_PATH = ROOT / "index.html"


def cell_time(v):
    if v is None:
        return None
    if isinstance(v, dt.time):
        return v.strftime("%H:%M")
    if isinstance(v, dt.datetime):
        return v.strftime("%H:%M")
    return None


def read_block(ws, r0, r1):
    items = []
    for r in range(r0, r1 + 1):
        seq = ws.cell(r, 1).value
        title = ws.cell(r, 2).value
        t = ws.cell(r, 3).value
        if seq is None or not str(seq).strip().isdigit():
            continue
        if title is None or not str(title).strip():
            continue
        hm = cell_time(t)
        if hm is None:
            continue
        items.append(
            {
                "serial": int(seq),
                "title": str(title).strip(),
                "deadline": hm,
            }
        )
    return items


def next_version():
    opath = ROOT / "docs" / "checklist-data.json"
    if not opath.exists():
        return 1
    try:
        old = json.loads(opath.read_text(encoding="utf-8"))
        v = old.get("version")
        return int(v) + 1 if isinstance(v, int) else 1
    except Exception:
        return 1


def embed_checklist_json_in_index(embedded_min_json: str) -> bool:
    """Replace #checklist-embedded-data script body in index.html. Returns True if updated."""
    if not HTML_PATH.exists():
        return False
    html = HTML_PATH.read_text(encoding="utf-8")
    if 'id="checklist-embedded-data"' not in html:
        print("index.html has no checklist-embedded-data; run scripts/patch_checklist_index.py once on a fresh file.")
        return False

    def repl(m):
        return m.group(1) + embedded_min_json + m.group(3)

    new_html, n = re.subn(
        r'(<script[^>]*\bid="checklist-embedded-data"[^>]*>)([\s\S]*?)(</script>)',
        repl,
        html,
        count=1,
    )
    if n != 1:
        print("embed replace failed (pattern not found or ambiguous)")
        return False
    HTML_PATH.write_text(new_html, encoding="utf-8")
    return True


def main():
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    day = read_block(ws, 3, 17)
    night = read_block(ws, 24, 32)
    dawn = read_block(ws, 39, 46)
    wb.close()
    out = {
        "version": next_version(),
        "source": "气象服务席检查单工作项.xlsx",
        "exportedAt": dt.datetime.now().replace(microsecond=0).isoformat(),
        "shifts": {
            "day": {"id": "day", "label": "每日 08:30–18:00", "items": day},
            "night": {"id": "night", "label": "每日 17:30–次日 03:30", "items": night},
            "dawn": {"id": "dawn", "label": "每日 03:00–09:00", "items": dawn},
        },
    }
    opath = ROOT / "docs" / "checklist-data.json"
    opath.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    emin = ROOT / "docs" / "checklist-embed.min.json"
    min_json = json.dumps(out, ensure_ascii=False, separators=(",", ":"))
    emin.write_text(min_json, encoding="utf-8")
    print("day", len(day), "night", len(night), "dawn", len(dawn))
    print("written", opath)
    print("written", emin)
    if embed_checklist_json_in_index(min_json):
        print("updated", HTML_PATH, "embedded checklist (hard-refresh browser)")
    else:
        print("index.html embed not updated")


if __name__ == "__main__":
    main()
