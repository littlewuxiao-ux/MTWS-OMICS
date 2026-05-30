
import re
import pandas as pd
from flask import Flask, request, jsonify, send_from_directory
from datetime import datetime, timedelta, timezone
import traceback
import os
import sys
import time
import threading
import copy
import json

# ==========================================
# [系统保护] Mock 类与安全导入
# ==========================================
class MockSFClient:
    def get_qrcode(self): return {"success": False, "error": "后端模块缺失(Mock)"}
    def check_scan_status(self): return {"success": False, "error": "后端模块缺失(Mock)"}
    def validate_login_and_get_token(self, t, s): return {"success": False, "message": "后端模块缺失(Mock)"}
    def get_session_status(self): return {"logged_in": False, "userCode": "OFFLINE"}
    def logout(self): pass
    def fetch_weather_data(self, *args): return ""

sf_client_instance = MockSFClient()
METARParser = None
parse_tafs = None

try:
    from .logic.metar_parser import METARParser
    from .logic.taf_parser import parse_tafs 
    from .logic.sf_client import sf_client_instance
except ImportError:
    try:
        from metar_parser import METARParser
        from taf_parser import parse_tafs
        from sf_client import sf_client_instance
    except ImportError:
        if METARParser is None:
            class METARParser:
                def parse(self, text, for_scoring=False): return {}
                def get_weather_severity(self, w): return 0
        if parse_tafs is None:
            def parse_tafs(text): return []

def get_base_path():
    if getattr(sys, 'frozen', False): return sys._MEIPASS
    else: return os.path.dirname(os.path.abspath(__file__))

base_path = get_base_path()

if os.path.exists(os.path.join(base_path, 'frontend')):
    frontend_folder = os.path.join(base_path, 'frontend')
elif os.path.exists(os.path.join(os.path.dirname(base_path), 'frontend')):
    frontend_folder = os.path.join(os.path.dirname(base_path), 'frontend')
else:
    frontend_folder = base_path

app = Flask(__name__, static_folder=frontend_folder)

# --- 常量定义 ---
CHINESE_WEATHER_MAP = {
    "雷雨": "TSRA", "强雷雨": "+TSRA", "大雷雨": "+TSRA", "弱雷雨": "-TSRA", "小雷雨": "-TSRA", "雷暴": "TS",
    "小雨": "-RA", "弱雨": "-RA", "雨": "RA", "中雨": "RA", "大雨": "+RA", "强雨": "+RA", 
    "小雪": "-SN", "弱雪": "-SN", "雪": "SN", "中雪": "SN", "大雪": "+SN", "强雪": "+SN",
    "小雨夹雪": "-RASN", "弱雨夹雪": "-RASN", "雨夹雪": "RASN",
    "冻雨": "FZRA", "小阵雨": "-SHRA", "弱阵雨": "-SHRA", "阵雨": "SHRA", "大阵雨": "+SHRA", "强阵雨": "+SHRA",
    "小阵雪": "-SHSN", "弱阵雪": "-SHSN", "阵雪": "SHSN", "大阵雪": "+SHSN", "强阵雪": "+SHSN",
    "冰雹": "GR", "小冰雹": "GS", "雾": "FG", "轻雾": "BR", "冻雾": "FZFG", "霾": "HZ", "烟": "FU",
    "沙尘暴": "SS", "扬沙": "SA", "尘": "DU", "晴": "SKC", "晴空": "SKC", "无天气": "NSW",
    "小毛毛雨": "-DZ", "弱毛毛雨": "-DZ", "毛毛雨": "DZ",
    "米雪": "SG", "小米雪": "-SG", 
    "低吹": "DS", "高吹": "BL", "龙卷": "FC", "飑": "SQ",
    "低吹雪": "DRSN", "高吹雪": "BLSN" 
}
METAR_TO_CHINESE_MAP = {v: k for k, v in CHINESE_WEATHER_MAP.items()}
METAR_TO_CHINESE_MAP.update({
    "+RA": "大雨", "-RA": "小雨", "RA": "中雨",
    "+SN": "大雪", "-SN": "小雪", "SN": "中雪",
    "-SG": "小米雪", "SG": "米雪",
    "+TSRA": "强雷雨", "-TSRA": "弱雷雨", "TSRA": "中雷雨",
    "+SHRA": "大阵雨", "-SHRA": "小阵雨", "SHRA": "中阵雨",
    "TS": "雷暴", "-RASN": "小雨夹雪", "-DZ": "小毛毛雨"
})
AIRPORT_NAME_MAP = {"ZBAA": "首都机场", "ZBAD": "大兴机场", "ZBTJ": "天津机场", "ZBSJ": "石家庄机场", "ZGSZ": "深圳机场", "ZSHC": "杭州机场", "ZHEC": "鄂州机场", "ZWWW": "乌鲁木齐"}
WIND_DIRECTION_MAP = {'N': 0, 'NNE': 22.5, 'NE': 45, 'ENE': 67.5, 'E': 90, 'ESE': 112.5, 'SE': 135, 'SSE': 157.5, 'S': 180, 'SSW': 202.5, 'SW': 225, 'WSW': 247.5, 'W': 270, 'WNW': 292.5, 'NW': 315, 'NNW': 337.5, 'VRB': 'VRB'}
CHINESE_WIND_DIR_MAP = {'N': '北风', 'NNE': '东北偏北风', 'NE': '东北风', 'ENE': '东北偏东风', 'E': '东风', 'ESE': '东南偏东风', 'SE': '东南风', 'SSE': '东南偏南风', 'S': '南风', 'SSW': '西南偏南风', 'SW': '西南风', 'WSW': '西南偏西风', 'W': '西风', 'WNW': '西北偏西风', 'NW': '西北风', 'NNW': '西北偏北风', 'VRB': '风向不定'}

@app.route('/')
def serve_index(): return send_from_directory(frontend_folder, 'index.html')
@app.route('/<path:path>')
def serve_static(path): return send_from_directory(frontend_folder, path)

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

    
# --- 辅助函数 ---
def _format_hourly_forecast_to_chinese(hour_data):
    if not hour_data: return "晴好"
    data_to_read = hour_data.get('base', hour_data)
    if hour_data.get('rule') in ['TEMPO', 'BECMG_TRANSITION']: data_to_read = {**hour_data.get('base', {}), **hour_data.get('change', {})}
    if not data_to_read: return "晴好"
    parts = []
    wind_dir = data_to_read.get('wind_dir'); wind_speed = data_to_read.get('wind_speed')
    if wind_dir and wind_speed and wind_speed > 0: parts.append(f"{CHINESE_WIND_DIR_MAP.get(wind_dir, wind_dir)}{wind_speed}米/秒")
    elif wind_speed and wind_speed > 0: parts.append(f"风速{wind_speed}米/秒")
    vis = data_to_read.get('visibility')
    if vis and vis < 9999: parts.append(f"能见度{vis}米")
    cloud = data_to_read.get('cloud')
    if cloud and cloud.get('height') is not None:
        amt = cloud.get('amount'); amt_cn = { 'BKN/OVC': '5-8量', 'FEW/SCT': '1-4量' }.get(amt, amt)
        parts.append(f"云{amt_cn} {cloud.get('height')}米")
    weather = data_to_read.get('weather', 'NSW').upper()
    if weather and weather != 'NSW':
        parts.append(" ".join([METAR_TO_CHINESE_MAP.get(code, code) for code in weather.split()]))
    return " ".join(parts) if parts else "晴好"

def generate_forecast_summary(airport_code, hourly_forecasts, mode='manual'):
    if mode == 'taf' or not hourly_forecasts: return ""
    airport_name = AIRPORT_NAME_MAP.get(airport_code, airport_code)
    beijing_tz = timezone(timedelta(hours=8)); bjt_forecasts = []
    for utc_hour_str in sorted(hourly_forecasts.keys()):
        if utc_hour_str.startswith('__'): continue # 🌟 修复：过滤掉非时间字符串的隐藏属性
        try:
            now = datetime.now(); day = int(utc_hour_str[:2]); hour = int(utc_hour_str[2:])
            utc_dt = datetime(now.year, now.month, day, hour, tzinfo=timezone.utc)
            bjt_dt = utc_dt.astimezone(beijing_tz)
            summary = _format_hourly_forecast_to_chinese(hourly_forecasts.get(utc_hour_str, {}))
            bjt_forecasts.append({"day": bjt_dt.day, "hour": bjt_dt.hour, "summary": summary})
        except: continue
    if not bjt_forecasts: return ""
    parts = []; i = 0
    while i < len(bjt_forecasts):
        curr = bjt_forecasts[i]; txt = curr['summary']
        if txt == "晴好": i += 1; continue
        start_d = curr['day']; start_h = curr['hour']; end_h = start_h; j = i + 1
        while j < len(bjt_forecasts) and bjt_forecasts[j]['summary'] == txt and bjt_forecasts[j]['day'] == start_d: end_h = bjt_forecasts[j]['hour']; j += 1
        time_range = f"{start_h:02d}时" if start_h == end_h else f"{start_h:02d}-{end_h:02d}时"
        parts.append(f"{start_d}日{time_range}{txt}"); i = j
    return f"{airport_name}预报 (北京时): {', '.join(parts)}。" if parts else f"{airport_name}预报 (北京时): 预计天气适航。"

def parse_manual_forecast_text(text):
    if not text: return {}
    parsed = {'weather': 'NSW', 'visibility': 9999, 'wind_speed': 0, 'cloud': {}}
    text = text.upper(); temp = text
    
    # 1. 提取风向风速 (如 N10)
    w_match = re.search(r'([A-Z]{1,3})(\d+)', text)
    if w_match and w_match.group(1) in WIND_DIRECTION_MAP:
        parsed['wind_speed'] = int(w_match.group(2)); parsed['wind_dir'] = w_match.group(1); temp = temp.replace(w_match.group(0), '', 1)
    
    # 🌟 新增核心功能：识别纯数字的快捷云高 (0, 30, 60, 90, 120)
    # \b 表示单词边界，确保不会误抓到 1300(能见度) 里的 30
    naked_cld = re.search(r'\b(0|30|60|90|120)\b', temp)
    if naked_cld:
        # 默认视为 BKN/OVC 的低云云高
        parsed['cloud'] = {'amount': 'BKN/OVC', 'height': int(naked_cld.group(1))}
        # 从文本中剔除这个数字，防止后续被误认为其他要素
        temp = re.sub(r'\b(0|30|60|90|120)\b', '', temp, count=1)
        
    # 2. 提取天气现象
    sorted_keys = sorted(CHINESE_WEATHER_MAP.keys(), key=len, reverse=True)
    wx_parts = []
    for cn in sorted_keys:
        if cn in temp:
            wx_parts.append(CHINESE_WEATHER_MAP[cn]); temp = temp.replace(cn, '')
    wx_parts.extend(re.findall(r'([+\-]?(?:VC|MI|PR|BC|BL|DR|SH|TS|FZ)?[A-Z]{2,6})', text))
    if wx_parts: parsed['weather'] = " ".join(filter(None, wx_parts))
    
    # 3. 提取能见度
    vis = re.search(r'(\d{3,4})M?', text)
    if vis: parsed['visibility'] = int(vis.group(1))
    
    # 4. 提取标准云高 (如 OVC001) - 如果有标准写法，将覆盖快捷数字写法
    cld = re.search(r'(FEW|SCT|BKN|OVC)(\d{3})', text)
    if cld: parsed['cloud'] = {'amount': 'BKN/OVC' if cld.group(1) in ['BKN','OVC'] else 'FEW/SCT', 'height': round(int(cld.group(2))*30.48)}
    
    # 5. 提取标准风速 (如 10MPS)
    spd = re.search(r'(\d{2,3})(?:G(\d{2,3}))?MPS', text)
    if spd: parsed['wind_speed'] = max(int(spd.group(1)), int(spd.group(2) or 0))
    
    return parsed

def _format_single_forecast_part(fcst_dict):
    if not fcst_dict: return []
    parts = []
    if fcst_dict.get('visibility') is not None: parts.append(f"能见度: {fcst_dict['visibility']}m")
    if fcst_dict.get('wind_speed') is not None: parts.append(f"风速: {fcst_dict['wind_speed']}mps")
    c = fcst_dict.get('cloud', {})
    if c: parts.append(f"云: {c.get('raw_amount') or c.get('amount')} at {c.get('height')}m")
    elif 'cloud' in fcst_dict: parts.append("云: NSC/CAVOK")
    weather = fcst_dict.get('weather', 'NSW')
    if weather and weather != 'NSW': parts.append(f"天气: {weather}")
    return parts

def format_forecast_for_display(fcst_dict):
    if not fcst_dict: return "无预报"
    rule = fcst_dict.get('rule', 'NORMAL')
    base_dict = fcst_dict.get('base', fcst_dict) 
    base_parts = _format_single_forecast_part(base_dict)
    base_str = "\n".join(base_parts) if base_parts else "CAVOK / NSW"
    if rule == 'NORMAL': return base_str
    change_parts = _format_single_forecast_part(fcst_dict.get('change', {}))
    change_str = "\n".join(change_parts)
    if not change_str: return base_str
    if rule == 'TEMPO': return f"主报:\n{base_str}\n\nTEMPO:\n{change_str}"
    if rule == 'BECMG_TRANSITION':
        merged_dict = {**base_dict, **fcst_dict.get('change', {})}
        merged_parts = _format_single_forecast_part(merged_dict)
        merged_str = "\n".join(merged_parts) if merged_parts else "CAVOK / NSW"
        return f"BECMG:\n{merged_str}"
    return base_str 

def _parse_ddhhmm_to_datetime(time_str, context_date, is_metar=False):
    # 此函数专用于处理 DDHHMM 格式
    try:
        if len(time_str) < 4: raise ValueError(f"Invalid time format: {time_str}")
        day = int(time_str[0:2]); hour = int(time_str[2:4]); minute = int(time_str[4:6]) if len(time_str) >= 6 else 0
        time_delta = timedelta(days=0)
        if hour == 24: hour = 0; time_delta = timedelta(days=1)
        
        # 锚点日期 (Context Date)
        naive_context = datetime(context_date.year, context_date.month, context_date.day)
        candidates = []
        
        # 尝试上个月、本月、下个月
        for month_offset in [-1, 0, 1]:
            year = naive_context.year
            month = naive_context.month + month_offset
            if month < 1: year -= 1; month = 12
            if month > 12: year += 1; month = 1
            try:
                candidate = datetime(year, month, day, hour, minute) + time_delta
                candidates.append(candidate.replace(tzinfo=timezone.utc))
            except ValueError: continue
        
        if not candidates: raise ValueError(f"无法为 {time_str} 创建有效日期")
        
        aware_context = context_date if context_date.tzinfo else context_date.replace(tzinfo=timezone.utc)
        # 找绝对距离最近的
        return min(candidates, key=lambda dt: abs(dt - aware_context))
            
    except Exception as e: raise ValueError(f"Time parsing error: {e}")

def _fill_forecast_gaps(hourly_forecasts):
    if not hourly_forecasts: return hourly_forecasts
    # 🌟 修复：只提取纯数字的时间键进行排序和补全，放过 '__' 开头的标记
    sorted_keys = sorted([k for k in hourly_forecasts.keys() if not k.startswith('__')])
    filled_forecasts = copy.deepcopy(hourly_forecasts)
    def next_ddhh(ddhh_str):
        d = int(ddhh_str[:2]); h = int(ddhh_str[2:])
        h += 1
        if h >= 24: h = 0; d += 1 
        return f"{d:02d}{h:02d}"
    for i in range(len(sorted_keys) - 1):
        curr_key = sorted_keys[i]; next_exist_key = sorted_keys[i+1]
        cursor = next_ddhh(curr_key)
        while cursor != next_exist_key:
            if len(cursor) != 4: break 
            base_data_to_copy = filled_forecasts.get(cursor, filled_forecasts[sorted_keys[i]])
            filled_forecasts[cursor] = copy.deepcopy(base_data_to_copy)
            cursor = next_ddhh(cursor)
            # 简单防止死循环
            if int(cursor[:2]) != int(curr_key[:2]) and abs(int(cursor[:2]) - int(curr_key[:2])) > 1 and int(curr_key[:2])!=31: break 
    return filled_forecasts

def parse_taf_string_robust(taf_text):
    clean_text = taf_text.strip().replace('=', '').replace('\n', ' ')
    clean_text = re.sub(r'\s+', ' ', clean_text)
    clean_text = clean_text.replace('TMEPO', 'TEMPO').replace('BEMCG', 'BECMG')
    validity_match = re.search(r'\b(\d{4}/\d{4})\b', clean_text)
    if not validity_match: return {}
    validity_str = validity_match.group(1)
    start_d = int(validity_str[:2]); start_h = int(validity_str[2:4]); end_d = int(validity_str[5:7]); end_h = int(validity_str[7:])
    tokens = clean_text.split()
    groups = []
    current_group = {'type': 'BASE', 'content': []}
    for token in tokens:
        if token in ['BECMG', 'TEMPO', 'FM']:
            groups.append(current_group); current_group = {'type': token, 'content': []}
        else: current_group['content'].append(token)
    groups.append(current_group)
    def parse_elements(tokens):
        data = {'wind_speed': None, 'visibility': None, 'weather': None, 'cloud': None}
        for t in tokens:
            if re.match(r'^\d{3}\d{2}(?:G\d{2})?MPS$', t) or re.match(r'^\d{3}\d{2}(?:G\d{2})?KT$', t): data['wind_speed'] = int(t[3:5])
            if re.match(r'^\d{4}$', t) and t != '9999': data['visibility'] = int(t)
            if t == '9999': data['visibility'] = 9999
            if re.search(r'(?:^|\+|-)(?:TS|SH|FZ|BL|DR|MI|BC|PR|RA|SN|SG|PL|GR|GS|DZ|FG|BR|HZ|FU|SA|DU|SS)(?:RA|SN|SG|PL|GR|GS|DZ)?$', t):
                if not re.search(r'\d', t): 
                    if data['weather']: data['weather'] += ' ' + t
                    else: data['weather'] = t
            if t == 'NSW': data['weather'] = 'NSW'
        return data
    base_group = groups[0]
    base_content_start_idx = 0
    for i, t in enumerate(base_group['content']):
        if re.match(r'\d{4}/\d{4}', t): base_content_start_idx = i + 1; break
    current_base = parse_elements(base_group['content'][base_content_start_idx:])
    hours = []
    curr_d, curr_h = start_d, start_h
    limit = 0
    while limit < 48:
        h_str = f"{curr_d:02d}{curr_h:02d}"
        hours.append(h_str)
        if curr_d == end_d and curr_h == end_h: break
        curr_h += 1
        if curr_h == 24: curr_h = 0; curr_d += 1
        limit += 1
    timeline = {h: {'base': copy.deepcopy(current_base), 'rule': 'NORMAL', 'change': {}} for h in hours}
    for grp in groups[1:]:
        g_type = grp['type']; content = grp['content']
        time_match = re.match(r'(\d{2})(\d{2})/(\d{2})(\d{2})', content[0]) if content else None
        if not time_match: continue
        changes = parse_elements(content[1:])
        sd, sh, ed, eh = int(time_match.group(1)), int(time_match.group(2)), int(time_match.group(3)), int(time_match.group(4))
        target_hours = []
        in_range = False
        for h_str in hours:
            hd, hh = int(h_str[:2]), int(h_str[2:])
            if hd == sd and hh == sh: in_range = True
            if hd == ed and hh == eh: in_range = False
            if in_range: target_hours.append(h_str)
        if g_type == 'BECMG':
            apply_from_idx = -1
            for idx, h in enumerate(hours):
                hd, hh = int(h[:2]), int(h[2:])
                if hd == ed and hh == eh: apply_from_idx = idx; break
            if apply_from_idx != -1:
                for i in range(apply_from_idx, len(hours)):
                    h_key = hours[i]; base_ref = timeline[h_key]['base']
                    if changes['wind_speed'] is not None: base_ref['wind_speed'] = changes['wind_speed']
                    if changes['visibility'] is not None: base_ref['visibility'] = changes['visibility']
                    if changes['weather'] is not None: base_ref['weather'] = changes['weather']
        elif g_type == 'TEMPO':
            for h_key in target_hours: timeline[h_key]['rule'] = 'TEMPO'; timeline[h_key]['change'] = changes
    return timeline

class QualityAssessor:
    def __init__(self, standards, custom_phenomena=None):
        self.standards = standards
        self.vis_map = { 'takeoff': standards.get('vis_takeoff', 400), 'landing': standards.get('vis_landing', 800), 'warning': standards.get('vis_warning', 1600) }
        self.cld_map = { 'takeoff': standards.get('cld_takeoff', 30), 'landing': standards.get('cld_landing', 60), 'warning': standards.get('cld_warning', 120) }
        if custom_phenomena: self.phen_categories = custom_phenomena
        else: self.phen_categories = { '雷雨类': ['TSRA'], '积冰类': ['FZDZ', 'FZRA', 'SN', 'SG', 'PL'], '强降水(无雷)类': ['RA', 'SH', 'SHRA'], '特殊类': ['GR', 'GS', 'FC', 'SQ'] }
        self.valid_codes_whitelist = set()
        for cats in self.phen_categories.values():
            for c in cats: self.valid_codes_whitelist.add(c)

    def _get_intensity(self, code):
        if not code or code == 'NSW': return '无'
        if '+' in code: return '强'
        if '-' in code: return '弱'
        return '中'

    def score_wind(self, fcst, obs):
        if fcst is None and obs is None: return '不评'
        f_spd = fcst or 0; o_spd = obs or 0
        if f_spd < 17 and o_spd < 17: return '不评' 
        if f_spd >= 17 and o_spd < 17: return '空报'
        if f_spd < 17 and o_spd >= 17: return '漏报'
        return '完美'

    def score_visibility(self, fcst, obs, obs_weather='NSW', fcst_weather='NSW'):
        f = fcst if fcst is not None else 9999
        o = obs if obs is not None else 9999
        def get_vis_level(v):
            if v <= self.vis_map['takeoff']: return 3
            if v <= self.vis_map['landing']: return 2
            if v <= self.vis_map['warning']: return 1
            return 0
        f_lvl = get_vis_level(f); o_lvl = get_vis_level(o)
        matrix = {
            3: {3: '完美', 2: '优秀', 1: '空报', 0: '空报'},
            2: {3: '优秀', 2: '完美', 1: '空报', 0: '空报'},
            1: {3: '漏报', 2: '漏报', 1: '完美', 0: '空报'},
            0: {3: '漏报', 2: '漏报', 1: '漏报', 0: '不评'} 
        }
        return matrix[f_lvl][o_lvl]

    def score_low_cloud(self, fcst_cloud, obs_cloud):
        f_h = fcst_cloud.get('height', 9999) if fcst_cloud else 9999
        f_amt_code = fcst_cloud.get('amount', '') if fcst_cloud else ''
        o_h = obs_cloud.get('height', 9999) if obs_cloud else 9999
        o_amt_code = obs_cloud.get('amount_code', '') if obs_cloud else ''
        def get_amt_level(code):
            if code in ['BKN', 'OVC', 'BKN/OVC']: return 2
            if code == 'SCT' or code == 'FEW/SCT': return 1
            return 0
        f_amt_lvl = get_amt_level(f_amt_code); o_amt_lvl = get_amt_level(o_amt_code); warn_h = self.cld_map['warning']
        if o_amt_lvl == 1 and o_h <= warn_h:
            if f_h <= warn_h: return '优秀'
            return '不评'
        if o_amt_lvl == 2 and o_h <= warn_h:
            if f_amt_lvl < 2: return '漏报'
            def get_h_level(h):
                if h <= self.cld_map['takeoff']: return 3
                if h <= self.cld_map['landing']: return 2
                if h <= self.cld_map['warning']: return 1
                return 0
            fh_lvl = get_h_level(f_h); oh_lvl = get_h_level(o_h)
            matrix = {
                3: {3: '完美', 2: '优秀', 1: '空报', 0: '空报'},
                2: {3: '优秀', 2: '完美', 1: '空报', 0: '空报'},
                1: {3: '漏报', 2: '漏报', 1: '完美', 0: '空报'},
                0: {3: '漏报', 2: '漏报', 1: '漏报', 0: '不评'}
            }
            return matrix[fh_lvl][oh_lvl]
        if (o_amt_lvl < 2 or o_h > warn_h):
            if f_amt_lvl == 2 and f_h <= warn_h: return '空报'
        return '不评'

    def score_single_weather_category(self, fcst_code, obs_code, category):
        has_fcst = fcst_code != 'NSW'; has_obs = obs_code != 'NSW'
        if not has_fcst and not has_obs: return '不评' 
        if has_fcst and not has_obs: return '空报'
        if not has_fcst and has_obs: return '漏报'
        f_int = self._get_intensity(fcst_code); o_int = self._get_intensity(obs_code)
        if f_int == o_int: return '完美'
        return '优秀' 

def run_evaluation(forecasts, final_obs, standards, obs_reports_raw, custom_phenomena=None, sorted_hours_list=None, ap_code="UNKNOWN", custom_thresholds=None):
    # 🌟 修复：取样本键时，跳过隐藏属性
    sample_key = next((k for k in forecasts.keys() if not k.startswith('__')), None) if forecasts else None
    if sample_key and 'base' not in forecasts[sample_key]: forecasts = _fill_forecast_gaps(forecasts)
    
    assessor = QualityAssessor(standards, custom_phenomena)
    score_results, obs_results = [], []
    
    # 🌟 修复：如果没传时间列表，自己提取时也要过滤掉隐藏属性
    hours_to_evaluate = sorted_hours_list if sorted_hours_list else sorted([k for k in forecasts.keys() if not k.startswith('__')])
    if not hours_to_evaluate: return pd.DataFrame(), pd.DataFrame(), ""
    
    tempo_blocks = {} 
    current_block_id = None; last_change_content = None; block_map = {} 
    
    def get_change_signature(c): return json.dumps(c, sort_keys=True)

    for hour in hours_to_evaluate:
        fcst = forecasts.get(hour, {}); rule = fcst.get('rule', 'NORMAL'); change = fcst.get('change', {})
        if rule == 'TEMPO':
            if not any(k in change for k in ['weather', 'visibility', 'wind_speed']): rule = 'NORMAL' 
            else:
                sig = get_change_signature(change)
                if sig != last_change_content:
                    current_block_id = hour; tempo_blocks[current_block_id] = {'hours': [], 'change': change}
                tempo_blocks[current_block_id]['hours'].append(hour); block_map[hour] = current_block_id; last_change_content = sig
        if rule != 'TEMPO': last_change_content = None 

    tempo_hits = {} 
    tempo_vis_hits = {} 

    def clean_weather_final(wx_str, current_category=None):
        if not wx_str or wx_str == 'NSW': return 'NSW'
        codes = wx_str.split(); cleaned = []
        for c in codes:
            if c.startswith('VC'): continue
            
            # 核心修改：如果是强降水类，且包含“-”(弱/小)，则忽略它
            if current_category == '强降水(无雷)类' and c.startswith('-'):
                continue 
                
            core_code = c.replace('+', '').replace('-', '')
            is_ts = 'TS' in core_code
            if core_code in assessor.valid_codes_whitelist or is_ts: cleaned.append(c)
        return " ".join(cleaned) if cleaned else 'NSW'

    def check_hit(change_dict, obs_dict, category=None):
        if category:
            codes = assessor.phen_categories.get(category, [])
            change_wx = change_dict.get('weather', 'NSW')
            obs_wx = obs_dict.get('weather', {}).get(category, 'NSW') 
            has_fcst_cat = any(c in change_wx for c in codes)
            if not has_fcst_cat: return False
            obs_clean = clean_weather_final(obs_wx, category); obs_codes = obs_clean.split()
            hit = any(any(oc.replace('+','').replace('-','') == c for c in codes) for oc in obs_codes)
            if 'TS' in codes and 'TS' in obs_clean: hit = True
            return hit
        return False 
    
    def check_vis_hit(change_vis, obs_vis_val):
        if change_vis is None: return False
        warn_std = standards.get('vis_warning', 1600)
        # 如果预测低能见度，且实况确实低
        if change_vis <= warn_std and obs_vis_val <= warn_std: return True
        return False

    for bid, bdata in tempo_blocks.items():
        tempo_hits[bid] = {}
        for cat in assessor.phen_categories.keys():
            is_hit = False
            for h in bdata['hours']:
                obs = final_obs.get(h, {})
                if check_hit(bdata['change'], obs, category=cat): is_hit = True; break
            tempo_hits[bid][cat] = is_hit
        
        is_vis_hit = False
        ch_vis = bdata['change'].get('visibility')
        if ch_vis is not None:
            for h in bdata['hours']:
                obs_v = final_obs.get(h, {}).get('visibility', 9999)
                if check_vis_hit(ch_vis, obs_v): is_vis_hit = True; break
        tempo_vis_hits[bid] = is_vis_hit

    for hour in hours_to_evaluate:
        obs = final_obs.get(hour, {}); fcst_data = forecasts.get(hour, {})
        score_row, obs_row = {'时次': hour}, {'时次': hour}
        rule = fcst_data.get('rule', 'NORMAL'); base_fcst = fcst_data.get('base', fcst_data); change_fcst = fcst_data.get('change', {})
        
        if rule == 'TEMPO' and hour not in block_map: rule = 'NORMAL'
        if rule != 'TEMPO': 
            rule = 'NORMAL' 
            if change_fcst: base_fcst = {**base_fcst, **change_fcst}

        # --- 1. 天气现象 ---
        for category, codes in assessor.phen_categories.items():
            obs_raw = obs.get('weather', {}).get(category, 'NSW'); obs_clean = clean_weather_final(obs_raw, category)
            obs_row[category] = obs_raw if obs_raw != 'NSW' else '/' 
            base_raw = base_fcst.get('weather', 'NSW'); base_clean = clean_weather_final(base_raw, category)
            base_phen = 'NSW'
            for p in base_clean.split():
                core_p = p.replace('+','').replace('-','')
                if core_p in codes or ('TS' in codes and 'TS' in core_p): base_phen = p; break
            has_base = base_phen != 'NSW'
            o_p = 'NSW'
            for p in obs_clean.split():
                core_p = p.replace('+','').replace('-','')
                if core_p in codes or ('TS' in codes and 'TS' in core_p): o_p = p; break
            has_obs = o_p != 'NSW'

            if rule == 'NORMAL':
                if not has_base and not has_obs: score_row[category] = '不评'
                elif has_base and not has_obs: score_row[category] = '空报'
                elif not has_base and has_obs: score_row[category] = '漏报'
                else: score_row[category] = assessor.score_single_weather_category(base_phen, o_p, category)
            elif rule == 'TEMPO':
                bid = block_map.get(hour); is_block_hit = tempo_hits.get(bid, {}).get(category, False)
                tempo_raw = change_fcst.get('weather', 'NSW'); tempo_clean = clean_weather_final(tempo_raw, category)
                tempo_phen = 'NSW'
                for p in tempo_clean.split():
                    core_p = p.replace('+','').replace('-','')
                    if core_p in codes or ('TS' in codes and 'TS' in core_p): tempo_phen = p; break
                has_tempo = tempo_phen != 'NSW'
                
                if not has_tempo:
                    if not has_base and not has_obs: score_row[category] = '不评'
                    elif has_base and not has_obs: score_row[category] = '空报' 
                    elif not has_base and has_obs: score_row[category] = '漏报' 
                    else: score_row[category] = assessor.score_single_weather_category(base_phen, o_p, category)
                    continue

                if is_block_hit:
                    if has_obs: score_row[category] = assessor.score_single_weather_category(tempo_phen, o_p, category)
                    else:
                        if not has_base: score_row[category] = '不评' 
                        else: score_row[category] = '空报' 
                else:
                    if has_obs: score_row[category] = '漏报' 
                    elif hour == tempo_blocks[bid]['hours'][0]: score_row[category] = '空报' 
                    else:
                        if not has_base and not has_obs: score_row[category] = '不评' 
                        elif has_base and not has_obs: score_row[category] = '空报'
                        elif not has_base and has_obs: score_row[category] = '漏报'
                        else: score_row[category] = assessor.score_single_weather_category(base_phen, o_p, category)

        # --- 2. 连续要素 (能见度) ---
        active_fcst_continuous = base_fcst
        
        # [诊断日志]
        current_vis_source = "Base"
        
        if rule == 'BECMG_TRANSITION' and change_fcst:
            active_fcst_continuous = {**base_fcst, **change_fcst}
            current_vis_source = "BECMG"
        
        elif rule == 'TEMPO':
            bid = block_map.get(hour)
            # 只有当 TEMPO 确实报了能见度变化时才判断
            if change_fcst.get('visibility') is not None:
                vis_hit = tempo_vis_hits.get(bid, False)
                if vis_hit:
                    # 命中：使用 TEMPO 值
                    active_fcst_continuous = {**base_fcst, **change_fcst}
                    current_vis_source = "TEMPO(Hit)"
                else:
                    # 未命中
                    if hour == tempo_blocks[bid]['hours'][0]:
                        # 首小时惩罚：强制用 TEMPO 值 (导致空报)
                        active_fcst_continuous = {**base_fcst, **change_fcst}
                        current_vis_source = "TEMPO(Miss-Penalty)"
                    else:
                        # 后续小时：回退 Base
                        active_fcst_continuous = base_fcst
                        current_vis_source = "TEMPO(Miss-Fallback)"
            else:
                # TEMPO 没报能见度，沿用 Base
                pass
        
        # 在黑框打印能见度取值过程
        vis_val_used = active_fcst_continuous.get('visibility', 9999)
        obs_val = obs.get('visibility', 9999)
        print(f"[DEBUG Vis] Time:{hour} | Rule:{rule} | Source:{current_vis_source} | Fcst:{vis_val_used} | Obs:{obs_val}")

        # === 🌟 全新大风打分逻辑 (对称 15% / 25%) ===
        # 1. 确定当前机场的大风阈值
        w_std = float(standards.get('wind_warning', 17))
        if custom_thresholds and ap_code in custom_thresholds:
            w_std = float(custom_thresholds[ap_code].get('ww', w_std))

        w_fcst_val = active_fcst_continuous.get('wind_speed')
        w_obs_val = obs.get('wind_speed')

        if w_fcst_val is None and w_obs_val is None:
            score_row['最大风速(MPS)'] = '不评'
        else:
            f_spd = float(w_fcst_val or 0)
            o_spd = float(w_obs_val or 0)

            if f_spd < w_std and o_spd < w_std:
                score_row['最大风速(MPS)'] = '不评'
            else:
                if o_spd == 0: # 防除零保护
                    if f_spd >= w_std: score_row['最大风速(MPS)'] = '空报'
                    else: score_row['最大风速(MPS)'] = '漏报'
                else:
                    # 核心公式
                    diff_ratio = (f_spd - o_spd) / o_spd
                    abs_ratio = abs(diff_ratio)

                    if abs_ratio < 0.15:
                        score_row['最大风速(MPS)'] = '完美'
                    elif abs_ratio < 0.25:
                        score_row['最大风速(MPS)'] = '优秀'
                    else:
                        if diff_ratio >= 0.25:
                            score_row['最大风速(MPS)'] = '空报'
                        else: # diff_ratio <= -0.25
                            score_row['最大风速(MPS)'] = '漏报'

        obs_row['最大风速(MPS)'] = w_obs_val
        
        vis_fcst = active_fcst_continuous.get('visibility'); vis_obs = obs.get('visibility')
        vis_wx_obs_raw = obs.get('weather', {})
        if isinstance(vis_wx_obs_raw, dict): vis_wx_obs = " ".join(vis_wx_obs_raw.values()) if vis_wx_obs_raw else 'NSW'
        else: vis_wx_obs = str(vis_wx_obs_raw)
        vis_wx_fcst = active_fcst_continuous.get('weather', 'NSW')
        
        score_row['最差能见度(m)'] = assessor.score_visibility(vis_fcst, vis_obs, vis_wx_obs, vis_wx_fcst)
        obs_row['最差能见度(m)'] = vis_obs

        cld_fcst = active_fcst_continuous.get('cloud'); cld_obs = obs.get('cloud')
        score_row['最低云高(m)'] = assessor.score_low_cloud(cld_fcst, cld_obs)
        obs_row['最低云高(m)'] = cld_obs.get('height') if cld_obs else '/'
        
        # 🌟 修复：使用本函数内的 forecasts 变量获取字典属性
        score_row['预报时效'] = forecasts.get('__validity', '24小时预报')
        score_row['预报全文'] = forecasts.get('__full_text', '-')
        score_row['预报内容'] = format_forecast_for_display(fcst_data).replace('\n', ' ')
        
        obs_row['预报时效'] = forecasts.get('__validity', '24小时预报')
        obs_row['预报全文'] = forecasts.get('__full_text', '-')
        obs_row['预报内容'] = '-'

        score_results.append(score_row); obs_results.append(obs_row)

    if not score_results: return pd.DataFrame(), pd.DataFrame(), ""
    df_scores = pd.DataFrame(score_results).set_index('时次'); df_obs = pd.DataFrame(obs_results).set_index('时次')
    
    # 🌟 修复核心1：在重置列名之前，先把“预报内容”单独保护起来！
    fcst_validity = df_scores['预报时效'] if '预报时效' in df_scores.columns else None
    fcst_full = df_scores['预报全文'] if '预报全文' in df_scores.columns else None
    fcst_content = df_scores['预报内容'] if '预报内容' in df_scores.columns else None

    obs_validity = df_obs['预报时效'] if '预报时效' in df_obs.columns else None
    obs_full = df_obs['预报全文'] if '预报全文' in df_obs.columns else None
    obs_content = df_obs['预报内容'] if '预报内容' in df_obs.columns else None

   # 🌟 恢复逻辑分组：基础要素一组，天气现象一组
    ordered_cols = ['最大风速(MPS)', '最差能见度(m)', '最低云高(m)', '雷雨类', '强降水(无雷)类', '积冰类', '特殊类']
    df_scores = df_scores.reindex(columns=ordered_cols).fillna('不评'); df_obs = df_obs.reindex(columns=ordered_cols).fillna('/')

    # 🌟 核心修复：按列（项目）独立统计，完美复刻你的专业计分逻辑！
    total_slots = len(df_scores)
    sum_perf = sum_exc = sum_fa = sum_miss = sum_eval = sum_acc = sum_tot = 0
    any_evaluated = False

    for col in ordered_cols:
        if col not in df_scores.columns:
            continue
        vals = df_scores[col].tolist()
        perf = vals.count('完美')
        exc = vals.count('优秀')
        fa = vals.count('空报')
        miss = vals.count('漏报')
        not_eval = vals.count('不评')
        
        eval_c = perf + exc + fa + miss
        # 只要该项目有任何参评记录，就把它全天时次拉入考核！
        if eval_c > 0:
            any_evaluated = True
            sum_perf += perf
            sum_exc += exc
            sum_fa += fa
            sum_miss += miss
            sum_eval += eval_c
            # 关键点：该项目下没犯错的“不评”时次，全部视为正确判断，算作“准确”！
            sum_acc += (perf + exc + not_eval) 
            sum_tot += total_slots

    # 兜底：如果一整天所有项目完全没有被激活（即全天适航无天气），总评和准确保底给满
    if not any_evaluated:
        sum_tot = total_slots
        sum_acc = total_slots
    
    stats_dict = {
        "完美": sum_perf,
        "优秀": sum_exc,
        "空报": sum_fa,
        "漏报": sum_miss,
        "准确": sum_acc,
        "参评": sum_eval,
        "总评": sum_tot
    }

    # 🌟 修复核心2：统计算完后，把“预报内容”插回第一列，传给前端和导出模块！
    if fcst_content is not None: df_scores.insert(0, '预报内容', fcst_content)
    if fcst_full is not None: df_scores.insert(0, '预报全文', fcst_full)
    if fcst_validity is not None: df_scores.insert(0, '预报时效', fcst_validity)
    
    if obs_content is not None: df_obs.insert(0, '预报内容', obs_content)
    if obs_full is not None: df_obs.insert(0, '预报全文', obs_full)
    if obs_validity is not None: df_obs.insert(0, '预报时效', obs_validity)

    return df_scores, df_obs, stats_dict    

# === 🌟 离线登录状态管理 ===
offline_session = {"logged_in": False, "userCode": "OFFLINE", "role": "user", "isOffline": True}

@app.route('/api/auth/offline_login', methods=['POST'])
def offline_login_api():
    global offline_session
    data = request.json
    username = data.get('username', '').strip() # 此时前端传过来的是【工号】
    password = data.get('password', '')

    if not username: return jsonify({"success": False, "message": "请输入工号"})

    # 密码默认依然是 000 (只是前端不再自动填入了，需要你手打)
    if password == "000":
        # 🌟 核心修改：认准工号 41060711，给予吴霄最高权限标识！
        is_admin = (username == "41060711")
        offline_session = {
            "logged_in": True,
            "userCode": username,
            "role": "admin" if is_admin else "user",
            "isOffline": True
        }
        return jsonify({"success": True, "message": f"登录成功！"})
    return jsonify({"success": False, "message": "密码错误"})

@app.route('/api/auth/qrcode', methods=['GET'])
def get_qrcode(): return jsonify(sf_client_instance.get_qrcode())

@app.route('/api/auth/check', methods=['GET'])
def check_qr_status(): return jsonify(sf_client_instance.check_scan_status())

@app.route('/api/auth/validate', methods=['POST'])
def validate_token():
    d = request.get_json()
    return jsonify(sf_client_instance.validate_login_and_get_token(d.get('ticket'), d.get('scan_id')))

@app.route('/api/auth/status', methods=['GET'])
def get_auth_status():
    # 优先检查离线登录状态
    if offline_session["logged_in"]:
        return jsonify(offline_session)
    # 否则返回实际的扫码状态，加上 isOffline 标记为 false
    status = sf_client_instance.get_session_status()
    status["isOffline"] = False
    # 如果扫码的是管理员账号，赋予 admin
    if status.get("userCode") == '41060711':
        status["role"] = "admin"
    return jsonify(status)

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    global offline_session
    offline_session = {"logged_in": False, "userCode": "OFFLINE", "role": "user", "isOffline": True}
    sf_client_instance.logout()
    return jsonify({"success": True})

@app.route('/api/fetch_data', methods=['POST'])
def fetch_weather_data():
    print(f"[DEBUG] 收到下载请求: {request.get_json()}")
    try:
        d = request.get_json()
        token = d.get('token'); start = d.get('start_time'); end = d.get('end_time'); aps = d.get('airports', ""); wtypes = d.get('wtypes', ["SA","SP"])
        now = datetime.now(timezone.utc)
        try:
            if '-' in start: fmt = "%Y-%m-%d"
            elif len(start) == 8: fmt = "%Y%m%d"
            elif len(start) == 12: fmt = "%Y%m%d%H%M"
            else: fmt = "%Y%m%d"
            s_dt = datetime.strptime(start, fmt).replace(tzinfo=timezone.utc)
            e_dt = datetime.strptime(end, fmt).replace(tzinfo=timezone.utc)
            if e_dt < s_dt: e_dt += timedelta(days=1)
            print(f"[DEBUG] 解析后的时间范围: {s_dt} -> {e_dt}")
        except ValueError as e: return jsonify({"success": False, "error": f"时间格式错误: {e}"}), 400
        
        t_aps = ["ZHEC", "ZBAA", "ZGHA", "ZSHC"]
        if aps: t_aps = [x.upper() for x in aps.strip().split() if x.strip()]
        data = sf_client_instance.fetch_weather_data(token, s_dt, e_dt, t_aps, wtypes)
        
        # 🌟 核心升级：直接调用后台现成的质量评定 TAF 引擎，将完美解析的逐时数据发给前端！
        parsed_tafs_list = []
        if "FT" in wtypes or "FC" in wtypes:
            try:
                parsed_tafs_list = parse_tafs(data, issued_date=s_dt.date())
                # 转换 datetime，避免 JSON 序列化报错
                for t in parsed_tafs_list:
                    if 'start_dt' in t and hasattr(t['start_dt'], 'isoformat'): t['start_dt'] = t['start_dt'].isoformat()
                    if 'end_dt' in t and hasattr(t['end_dt'], 'isoformat'): t['end_dt'] = t['end_dt'].isoformat()
            except Exception as e:
                print(f"TAF 预解析异常: {e}")

        print(f"[DEBUG] 丰台接口返回数据长度: {len(data) if data else 0}")
        return jsonify({"success": True, "data": data, "parsed_tafs": parsed_tafs_list})
    except Exception as e: return jsonify({"success": False, "error": str(e)}), 500
# ==========================================
# 🌟 新增：处理前端获取航班数据的请求
# ==========================================
@app.route('/api/fetch_flights', methods=['POST'])
def api_fetch_flights():
    data = request.json
    token = data.get('token')
    flight_date = data.get('flight_date')
    
    if not token or not flight_date:
        return jsonify({"success": False, "error": "缺少 token 或 flight_date 参数"})
        
    try:
        # 直接调用你已经在 sf_client_instance 里写好的方法
        flights = sf_client_instance.fetch_flight_schedule(token, flight_date)
        return jsonify({"success": True, "data": flights})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

# ==========================================
# 🌟 新增：接收前端机场字典，并直接覆盖保存到 airports.js
# ==========================================
@app.route('/api/save_airports', methods=['POST'])
def api_save_airports():
    print("\n" + "="*40)
    print("[DEBUG] 收到前端请求：准备保存新的机场字典配置...")
    
    data = request.json
    coords = data.get('coords', {})
    names = data.get('names', {})
    
    # 将字典格式化为标准的 JS 代码字符串
    import json
    js_content = f"window.AIRPORT_COORDS = {json.dumps(coords, ensure_ascii=False, indent=4)};\n\n"
    js_content += f"window.GLOBAL_AIRPORT_NAME_MAP = {json.dumps(names, ensure_ascii=False, indent=4)};\n"
    
    try:
        # 🌟 核心修复：直接使用 app.py 顶部已经定义好的 frontend_folder 全局路径
        global frontend_folder
        file_path = os.path.join(frontend_folder, 'airports.js')
        
        print(f"[DEBUG] 计算出的目标物理路径为: {file_path}")
        
        if not os.path.exists(file_path):
            print("[WARN] 警告：目标路径下未找到原 airports.js，系统将自动创建新文件。")
        else:
            print("[DEBUG] 检测到原 airports.js，准备执行数据覆写。")
            
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(js_content)
            
        print("[DEBUG] ✅ 写入成功！新的机场配置已保存。")
        print("="*40 + "\n")
        
        return jsonify({"success": True, "message": "已成功保存到物理文件！"})
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"[ERROR] ❌ 保存失败，详细报错信息如下:\n{error_msg}")
        return jsonify({"success": False, "error": str(e)})
        
@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    def kill():
        time.sleep(1)
        os._exit(0)
    threading.Thread(target=kill).start()
    return jsonify({"status": "shutting down"})

@app.route('/api/score', methods=['POST', 'OPTIONS']) 
def handle_scoring():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'})
    try:
        print("[DEBUG] 开始评分...")
        parser = METARParser(); d = request.get_json()
        mode = d.get('forecast_mode'); obs_text = d.get('obs_text', '')
        standards = {k: int(v) for k, v in d.get('standards', {}).items()}
        custom_thresholds = d.get('custom_thresholds', {})
        phenomena_config = d.get('phenomena_config', None)
        
        # 🌟 修复 1：去 export_config 里把前端藏好的日期挖出来！
        export_cfg = d.get('export_config', {})
        base_date_str = d.get('base_date_str') or export_cfg.get('base_date_str')
        
        issued_date = None
        if base_date_str:
            try:
                if '-' in base_date_str: 
                    issued_date = datetime.strptime(base_date_str, "%Y-%m-%d").date()
                else: 
                    issued_date = datetime.strptime(base_date_str[:8], "%Y%m%d").date()
            except Exception as e: 
                print(f"日期解析失败: {e}")

        # 🌟 修复 2：去掉“仅限席位预报”的束缚，全局接管前后 3 天的超大时间边界！
        s_str = d.get('start_time'); e_str = d.get('end_time')
        start_eval_dt = None; end_eval_dt = None

        if s_str and e_str:
            try:
                fmt_s = "%Y%m%d%H%M%S" if len(s_str) == 14 else ("%Y%m%d%H%M" if len(s_str) == 12 else "%Y%m%d")
                start_eval_dt = datetime.strptime(s_str, fmt_s).replace(tzinfo=timezone.utc)
                fmt_e = "%Y%m%d%H%M%S" if len(e_str) == 14 else ("%Y%m%d%H%M" if len(e_str) == 12 else "%Y%m%d")
                end_eval_dt = datetime.strptime(e_str, fmt_e).replace(tzinfo=timezone.utc)
            except Exception as e: 
                print(f"边界时间解析失败: {e}")

        raw_obs_list = []
        for line in obs_text.strip().split('\n'):
            if not line.strip(): continue
            parsed = parser.parse(line, for_scoring=True)
            if parsed and parsed.get('time'):
                raw_obs_list.append(parsed)

        results = {}
        assessor_cats = QualityAssessor(standards, phenomena_config)

        def aggregate_obs(reports):
            if not reports: return {}
            worst = min(reports, key=lambda r: (r.get('visibility', 9999), r.get('cloud_height', 9999) if r.get('cloud_height') is not None else 9999))
            wind = max([r.get('wind_speed', 0) for r in reports] or [0])
            bkn = [r for r in reports if r.get('cloud_amount') in ['BKN','OVC']]
            sct = [r for r in reports if r.get('cloud_amount') == 'SCT']
            cld_rep = min(bkn, key=lambda x:x['cloud_height']) if bkn else (min(sct, key=lambda x:x['cloud_height']) if sct else {})
            wx_cats = {}
            all_wx = [p for r in reports for p in (r.get('weather') or 'NSW').split()]
            for cat, codes in assessor_cats.phen_categories.items():
                rels = []
                for p in all_wx:
                    core_p = p.replace('+','').replace('-','')
                    if core_p in codes or ('TS' in codes and 'TS' in core_p): rels.append(p)
                if rels: wx_cats[cat] = max(rels, key=parser.get_weather_severity)
            return {'wind_speed': wind, 'visibility': worst.get('visibility', 9999), 'cloud': {'height': cld_rep.get('cloud_height'), 'amount_code': cld_rep.get('cloud_amount')}, 'weather': wx_cats}

        forecasts_map = {}
        forecasts_map = {}
        if mode == 'manual':
            raw_fcst = d.get('manual_forecasts', {})
            # 🌟 修复：直接获取前端传来的准确时效文本 (如: 未来24小时)
            time_range_text = d.get('time_range_text', '席位综合预报')
            for ap, hours_data in raw_fcst.items():
                fcsts = {h: parse_manual_forecast_text(txt) for h, txt in hours_data.items()}
                fcsts['__validity'] = time_range_text
                fcsts['__full_text'] = generate_forecast_summary(ap, fcsts, 'manual')
                forecasts_map[ap] = fcsts
        else:
            raw_taf_text = d.get('taf_text', '')
            parsed_tafs = parse_tafs(raw_taf_text, issued_date=issued_date)
            taf_blocks = re.split(r'(?=TAF(?:\s|AMD|COR))', raw_taf_text)
            
            # 🌟 需求2：智能提取与 AMD 排序机制
            recognize_amd = d.get('recognize_amd', False)
            temp_groups = {}
            for t in parsed_tafs:
                g_key = f"{t['airport']}_{t['validity_str']}"
                if g_key not in temp_groups: temp_groups[g_key] = []
                temp_groups[g_key].append(t)

            for group_key, t_list in temp_groups.items():
                if not recognize_amd:
                    # 🔴 模式A（关闭开关）：只考核真正的原报
                    original_t = None
                    original_raw = ""
                    for t in t_list:
                        # 提取这段报文对应的真实原始文本
                        raw_txt = ""
                        for b in taf_blocks:
                            if t['airport'] in b and t['validity_str'].replace('-', '/') in b:
                                raw_txt = b.strip(); break
                        
                        # 🔍 核心修复：精准定位！只有文本里没有 AMD 且没有 COR，才是真正的原报
                        if " AMD " not in raw_txt and " COR " not in raw_txt and not raw_txt.startswith("TAF AMD") and not raw_txt.startswith("TAF COR"):
                            original_t = t
                            original_raw = raw_txt
                            break
                            
                    if not original_t:
                        # ⚠️ 如果粘贴的这段时间里全是 AMD，没有原报（说明原报在昨天的文件里没粘过来）
                        # 那么直接跳过这组报文！保证不把 AMD 误当成原报评分！
                        continue
                        
                    fcsts = original_t['forecasts']
                    fcsts['__validity'] = original_t['validity_str']
                    fcsts['__full_text'] = original_raw if original_raw else f"TAF {original_t['airport']} {original_t['validity_str']}"
                    forecasts_map[group_key] = fcsts
                    
                else:
                    # 🟢 模式B（打开开关）：分离模式并智能标注
                    amd_counter = 1
                    for idx, t in enumerate(t_list):
                        matching_blocks = [b.strip() for b in taf_blocks if t['airport'] in b and t['validity_str'].replace('-', '/') in b]
                        raw_txt = matching_blocks[idx] if idx < len(matching_blocks) else (matching_blocks[-1] if matching_blocks else f"TAF {t['airport']} {t['validity_str']}")
                        
                        # 🔍 核心修复：通过文本内容定位，而不是依靠先后顺序
                        is_amd = " AMD " in raw_txt or " COR " in raw_txt or raw_txt.startswith("TAF AMD") or raw_txt.startswith("TAF COR")
                        
                        if not is_amd:
                            suffix = "" # 真正的原报，不加后缀
                        else:
                            suffix = f" (AMD-{amd_counter})"
                            amd_counter += 1 # 只有遇到 AMD 计数器才 +1
                            
                        unique_key = f"{group_key}{suffix}"
                        fcsts = t['forecasts']
                        fcsts['__validity'] = f"{t['validity_str']}{suffix}"
                        fcsts['__full_text'] = raw_txt
                        forecasts_map[unique_key] = fcsts

        for key, hourly_forecasts in forecasts_map.items():
            ap_code = key.split('_')[0]

            ap_obs = [r for r in raw_obs_list if r['station'] == ap_code]
            grouped_obs = {}
            for o in ap_obs:
                time_str = o['time'] 
                if len(time_str) >= 4:
                    ddhh = time_str[:4]
                    grouped_obs.setdefault(ddhh, []).append(o)
            
            final_obs_agg = {h: aggregate_obs(reps) for h, reps in grouped_obs.items()}
            
            sorted_hours = []
            if start_eval_dt and end_eval_dt:
                temp_list = []
                for h_str in hourly_forecasts.keys():
                    if h_str.startswith('__'): continue # 🌟 修复：过滤掉隐藏属性
                    try:
                        dt_val = _parse_ddhhmm_to_datetime(h_str + "00", start_eval_dt)
                        if start_eval_dt <= dt_val <= end_eval_dt:
                            temp_list.append((dt_val, h_str))
                    except Exception as e: 
                        pass
                
                temp_list.sort(key=lambda x: x[0])
                sorted_hours = [x[1] for x in temp_list]
            else:
                sorted_hours = [k for k in sorted(list(hourly_forecasts.keys())) if not k.startswith('__')] # 🌟 修复：过滤

            df_s, df_o, stats = run_evaluation(hourly_forecasts, final_obs_agg, standards, ap_obs, phenomena_config, sorted_hours_list=sorted_hours, ap_code=ap_code, custom_thresholds=custom_thresholds)
            
            if df_s.empty: continue
            res_entry = {'scores': df_s.reset_index().to_dict('records'), 'observations': df_o.reset_index().to_dict('records'), 'statistics': stats}
            if mode == 'taf':
                # 🌟 修复：过滤掉隐藏属性，只拿真实的时次字典去格式化
                res_entry['recognized_forecasts'] = [{"hour": h, "text": format_forecast_for_display(hourly_forecasts[h])} for h in sorted(hourly_forecasts.keys()) if not h.startswith('__')]
                res_entry['forecast_summary_bjt'] = generate_forecast_summary(ap_code, hourly_forecasts, 'taf')
            else:
                res_entry['forecast_summary_bjt'] = generate_forecast_summary(ap_code, hourly_forecasts, 'manual')
            results[key] = res_entry

        return jsonify({"success": True, "data": results})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

# 需求3：使用 Tkinter 原生弹出目录选择器
@app.route('/api/select_folder', methods=['GET'])
def select_folder_api():
    try:
        import tkinter as tk
        from tkinter import filedialog
        
        # 创建一个隐藏的 tk 根窗口
        root = tk.Tk()
        root.withdraw()
        # 强制窗口在最顶层
        root.attributes('-topmost', True)
        
        folder_path = filedialog.askdirectory(title="选择保存评定文件的目录")
        root.destroy()
        
        return jsonify({"success": True, "path": folder_path})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# 🌟 修复：获取当前电脑桌面绝对路径的接口（读取注册表穿透 D盘 / OneDrive）
@app.route('/api/get_desktop_path', methods=['GET'])
def get_desktop_path():
    import os
    import winreg
    desktop_dir = ""
    try:
        # 🌟 核心修复3：直接读取 Windows 注册表，无论桌面被移动到 D 盘还是哪里，都能精准定位！
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders")
        reg_path, _ = winreg.QueryValueEx(key, "Desktop")
        winreg.CloseKey(key)
        # 解析 %USERPROFILE% 等环境变量，得到最终真实绝对路径
        desktop_dir = os.path.expandvars(reg_path) 
    except Exception:
        # 备选方案：常规路径和 OneDrive 路径
        user_profile = os.path.expanduser("~")
        onedrive_desktop = os.path.join(user_profile, 'OneDrive', 'Desktop')
        if os.path.exists(onedrive_desktop):
            desktop_dir = onedrive_desktop
        else:
            desktop_dir = os.path.join(user_profile, 'Desktop')
    
    # 强制在桌面创建一个显眼的专属文件夹
    final_path = os.path.join(desktop_dir, 'SF预报评定导出')
    os.makedirs(final_path, exist_ok=True)
    return jsonify({"success": True, "path": final_path})

# === 新增：手动保存接口 (替换 payload 为 data 解决报错) ===
@app.route('/api/save_score', methods=['POST'])
def save_score_api():
    try:
        data = request.json
        results = data.get('results')
        mode = data.get('forecast_mode')
        export_config = data.get('export_config', {})
        
        backup_path = export_config.get('backup_path', '')
        # 🌟 核心修复：如果前端传来的是空字符串(普通人员或未配置)，强行兜底到程序目录的 backup 文件夹！
        if not backup_path: 
            backup_path = os.path.join(os.getcwd(), 'backup')
            
        excel_root = export_config.get('excel_root', '')
        eval_person = export_config.get('eval_person', '未指定人员')
        base_date_str = export_config.get('base_date_str', '')
        rater = export_config.get('rater', '未知') # 🌟 提取评分人
        
        # === 🌟 修复关键：兼容性导入必须放在调用之前 ===
        try:
            from logic.exporter import process_stats_and_save
        except ImportError:
            try:
                from .logic.exporter import process_stats_and_save
            except:
                from exporter import process_stats_and_save

        # 🌟 并且只调用一次，带有完整的 rater 参数
        process_stats_and_save(results, mode, backup_path, excel_root, eval_person, base_date_str, rater)
        
        return jsonify({"success": True, "message": "保存成功！"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500
# 🌟 找回失落的接口：日数据明细查询
@app.route('/api/query_raw_data', methods=['POST'])
def query_raw_data_api():
    try:
        data = request.json
        s_type = data.get('stats_type') 
        person = data.get('person')
        airport = data.get('airport', '').strip().upper()
        base_date = data.get('base_date')
        
        base_dates = [d.strip() for d in base_date.split(',')] if base_date else []
        if not base_dates: return jsonify({"success": False, "error": "请选择明细日期"})
        
        backup_path = data.get('backup_path', 'backup')
        if not backup_path: backup_path = os.path.join(os.getcwd(), 'backup')
        
        files_to_read = set()
        for d_str in base_dates:
            try:
                dt = datetime.strptime(d_str, '%Y-%m-%d')
                files_to_read.add(f"backup_{dt.strftime('%Y%m')}.json")
            except: pass
            
        raw_list = []
        for fname in files_to_read:
            fpath = os.path.join(backup_path, fname)
            if not os.path.exists(fpath): continue
                
            with open(fpath, 'r', encoding='utf-8') as f: db = json.load(f)
            
            for rec_id, rec in db.get('records', {}).items():
                if rec['mode'] != s_type: continue
                if s_type == 'manual' and person != 'ALL' and rec['person'] != person: continue
                if airport and rec['airport'] != airport: continue
                if rec['date'] not in base_dates: continue
                
                ap_code = rec['airport']
                scores = rec.get('scores', [])
                obs_dict = {o['时次']: o for o in rec.get('observations', [])}
                
                # 🌟 修复需求1：强制锁定 JSON 提取的顺序（基础气象要素在前，天气现象在后）
                ordered_items = ['最大风速(MPS)', '最差能见度(m)', '最低云高(m)', '雷雨类', '强降水(无雷)类', '积冰类', '特殊类']
                items = [k for k in ordered_items if scores and k in scores[0].keys()]
                
                for s_row in scores:
                    t = s_row['时次']
                    o_row = obs_dict.get(t, {})
                    row_data = {
                        "评定对象": rec['person'], 
                        "机场": ap_code, 
                        "预报时效": s_row.get('预报时效', '/'),
                        "时次(UTC)": t, 
                        "预报全文": s_row.get('预报全文', '/')
                    }
                    for item in items:
                        row_data[f"{item}(实况)"] = o_row.get(item, '/')
                        row_data[f"{item}(评价)"] = s_row.get(item, '/')
                    raw_list.append(row_data)
                    
        if not raw_list: return jsonify({"success": False, "error": "所选日期未找到明细数据！"})
        return jsonify({"success": True, "data": raw_list})
    except Exception as e: return jsonify({"success": False, "error": str(e)})
    
@app.route('/api/query_stats', methods=['POST'])
def query_stats_api():
    try:
        data = request.json
        s_type = data.get('stats_type') 
        person = data.get('person')
        airport = data.get('airport', '').strip().upper()
        time_type = data.get('time_type') 
        base_date = data.get('base_date')
        
        # 接收逗号分隔的多日数据
        base_dates = [d.strip() for d in base_date.split(',')] if base_date else []
        if not base_dates: return jsonify({"success": False, "error": "请选择评定日期"})
        
        backup_path = data.get('backup_path', 'backup')
        if not backup_path: backup_path = os.path.join(os.getcwd(), 'backup')

        files_to_read = set()
        if time_type == 'day':
            for d_str in base_dates:
                dt = datetime.strptime(d_str, '%Y-%m-%d')
                files_to_read.add(f"backup_{dt.strftime('%Y%m')}.json")
        elif time_type in ['month', 'year']:
            dt = datetime.strptime(base_dates[0], '%Y-%m-%d')
            if time_type == 'month': files_to_read.add(f"backup_{dt.strftime('%Y%m')}.json")
            else:
                for m in range(1, 13): files_to_read.add(f"backup_{dt.year}{m:02d}.json")

        aggregated_stats = {}
        for fname in files_to_read:
            fpath = os.path.join(backup_path, fname)
            if not os.path.exists(fpath): continue
            with open(fpath, 'r', encoding='utf-8') as f: db = json.load(f)
            
            for rec_id, rec in db.get('records', {}).items():
                # 🌟 核心拦截：绝对禁止机场预报和席位预报的数据互相串门污染！
                if rec.get('mode') != s_type: continue 
                
                if s_type == 'manual' and person != 'ALL' and rec['person'] != person: continue
                if airport and rec['airport'] != airport: continue
                
                try:
                    rec_dt = datetime.strptime(rec['date'], '%Y-%m-%d')
                except: continue

                # 🌟 按不同时间颗粒度精确放行
                if time_type == 'day' and rec['date'] not in base_dates: continue
                if time_type == 'month' and (rec_dt.year != dt.year or rec_dt.month != dt.month): continue
                if time_type == 'year' and rec_dt.year != dt.year: continue
                
                group_key = rec['airport'] if s_type == 'taf' else rec['person']
                if group_key not in aggregated_stats:
                    aggregated_stats[group_key] = {"总评":0, "参评":0, "准确":0, "空报":0, "漏报":0, "完美":0, "优秀":0, "适航天数":0}
                
                if rec.get('is_shihang'):
                    aggregated_stats[group_key]["适航天数"] += 1
                    slots = len(rec.get('scores', []))
                    if slots == 0: slots = 24 
                    aggregated_stats[group_key]["总评"] += slots
                    aggregated_stats[group_key]["准确"] += slots
                else:
                    for cat, st in rec['daily_stats'].items():
                        aggregated_stats[group_key]["总评"] += st.get('总评', 0)
                        aggregated_stats[group_key]["参评"] += st.get('参评', 0)
                        aggregated_stats[group_key]["准确"] += st.get('准确', 0)
                        aggregated_stats[group_key]["空报"] += st.get('空报', 0)
                        aggregated_stats[group_key]["漏报"] += st.get('漏报', 0)
                        aggregated_stats[group_key]["完美"] += st.get('完美', 0)
                        aggregated_stats[group_key]["优秀"] += st.get('优秀', 0)

        results = []
        for key, st in aggregated_stats.items():
            tot = st["总评"]; eval_c = st["参评"]; acc_c = st["准确"]; perf = st["完美"]; exc = st["优秀"]
            results.append({
                "统计对象": key, "总评项次": tot, "准确项次": acc_c, "参评项次": eval_c,
                "完美项次": perf, "优秀项次": exc, "空报项次": st["空报"], "漏报项次": st["漏报"],
                "准确率": f"{(acc_c/tot*100):.2f}%" if tot > 0 else "/",
                "完美率": f"{(perf/eval_c*100):.2f}%" if eval_c > 0 else "/",
                "优秀率": f"{(exc/eval_c*100):.2f}%" if eval_c > 0 else "/",
                "空报率": f"{(st['空报']/tot*100):.2f}%" if tot > 0 else "/",
                "漏报率": f"{(st['漏报']/tot*100):.2f}%" if tot > 0 else "/"
            })

        if not results: return jsonify({"success": False, "error": "未找到历史数据！"})
        return jsonify({"success": True, "data": results, "time_type": time_type, "base_date": base_date})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

# 🌟 从云端 Excel 逆向扫描重建 JSON 微型数据库的核心引擎 (智能识别路径与表头修复版)
@app.route('/api/sync_excel', methods=['POST'])
def sync_excel_api():
    try:
        data = request.json
        excel_root = data.get('excel_root')
        backup_path = data.get('backup_path')
        if not backup_path: backup_path = os.path.join(os.getcwd(), 'backup')
        
        if not excel_root or not os.path.exists(excel_root):
            return jsonify({"success": False, "error": "Excel 云盘目录不存在或未配置，无法同步！"})
            
        os.makedirs(backup_path, exist_ok=True)
        db_cache = {}
        
        for root, dirs, files in os.walk(excel_root):
            for file in files:
                if not file.endswith('.xlsx') or file.startswith('~'): continue
                filepath = os.path.join(root, file)
                
                # 🌟 修复核心 1：不依赖表格内部文字，直接从文件路径和名称中“侦探式”提取精准的年月日！
                try:
                    # 1. 从路径提取年份 (例如 ".../2026年/...")
                    year_match = re.search(r'(\d{4})年', root)
                    if not year_match: continue
                    year = int(year_match.group(1))
                    
                    # 2. 从文件名提取月日 (例如 席位预报质量评价表-曹骏-0415.xlsx -> 0415)
                    md_match = re.search(r'-(\d{4})\.xlsx$', file)
                    if not md_match: continue
                    mmdd = md_match.group(1)
                    month = int(mmdd[:2])
                    day = int(mmdd[2:])
                    
                    base_date_str = f"{year}-{month:02d}-{day:02d}"
                    dt = datetime(year, month, day)
                    month_key = dt.strftime('%Y%m')
                except Exception as e:
                    continue # 命名不规范的废弃文件直接跳过
                    
                try:
                    # 🌟 修复核心 2：用 header=None 读取，避免 Pandas 把带有“评分人”的合并单元格当成错误表头
                    df_s1_raw = pd.read_excel(filepath, sheet_name='日统计', header=None)
                    df_s2 = pd.read_excel(filepath, sheet_name='要素分解与汇总')
                except: continue
                
                if df_s2.empty or df_s1_raw.empty: continue
                
                # 🌟 修复核心 3：智能往下寻找真实的表头行
                header_idx = 0
                for idx, row in df_s1_raw.iterrows():
                    if '天气分类' in [str(x) for x in row.values]:
                        header_idx = idx
                        break
                        
                # 重新定位表头，完美截取有效数据
                df_s1 = df_s1_raw.iloc[header_idx+1:].copy()
                df_s1.columns = df_s1_raw.iloc[header_idx]
                
                # 🌟 修复核心1：向下填充合并单元格，防止读取多个机场时名字变成 NaN
                if '机场' in df_s1.columns:
                    df_s1['机场'] = df_s1['机场'].ffill()
                
                if month_key not in db_cache: db_cache[month_key] = {"month": month_key, "records": {}}
                    
                mode = 'taf'
                person = 'ALL'
                if '席位' in file:
                    mode = 'manual'
                    p_match = re.search(r'-(.*?)-\d{4}\.xlsx', file)
                    if p_match: person = p_match.group(1).strip()
                
                # 🌟 修复核心2：按机场分组循环，完美保护每个机场的空报漏报不被后续机场覆盖！
                if '机场' not in df_s1.columns: continue
                for ap_code, ap_df in df_s1.groupby('机场'):
                    if str(ap_code).startswith('评分人') or str(ap_code) == 'nan': continue
                    ap_code = str(ap_code).strip()
                    
                    rec_id = f"{base_date_str}_{mode}_{person}_{ap_code}"
                    daily_stats = {}
                    is_shihang = False
                    
                    # 录入每日统计 (Sheet 1)
                    for idx, row in ap_df.iterrows():
                        cat = row.get('天气分类'); stat_type = row.get('统计')
                        if pd.isna(cat): continue
                        if cat == '适航' and stat_type == '次数': is_shihang = True
                        if cat and cat not in ['总计', '适航'] and not str(cat).startswith('机场:'):
                            if stat_type == '次数':
                                daily_stats[cat] = {
                                    "完美": int(row.get('完美',0) if not pd.isna(row.get('完美',0)) else 0), 
                                    "优秀": int(row.get('优秀',0) if not pd.isna(row.get('优秀',0)) else 0), 
                                    "空报": int(row.get('空报',0) if not pd.isna(row.get('空报',0)) else 0), 
                                    "漏报": int(row.get('漏报',0) if not pd.isna(row.get('漏报',0)) else 0),
                                    "准确": int(row.get('准确',0) if not pd.isna(row.get('准确',0)) else 0), 
                                    "参评": int(row.get('参评',0) if not pd.isna(row.get('参评',0)) else 0), 
                                    "总评": int(row.get('总评',0) if not pd.isna(row.get('总评',0)) else 0)
                                }
                    
                    # 录入逐时详情 (Sheet 2)
                    ap_s2 = df_s2[df_s2['机场'] == ap_code] if '机场' in df_s2.columns else df_s2
                    scores, observations = [], []
                    for _, s2_row in ap_s2.iterrows():
                        time_str = str(s2_row.get('时次') or s2_row.get('时次(UTC)'))
                        if time_str == 'nan': continue
                        score_obj, obs_obj = {'时次': time_str}, {'时次': time_str}
                        
                        if '预报时效' in ap_s2.columns: score_obj['预报时效'] = str(s2_row['预报时效'])
                        if '预报全文' in ap_s2.columns: score_obj['预报全文'] = str(s2_row['预报全文'])
                        
                        for col in ap_s2.columns:
                            if '(实况)' in col: obs_obj[col.replace('(实况)', '').strip()] = s2_row[col]
                            elif '(评分)' in col or '(评价)' in col: score_obj[re.sub(r'\(评分\)|\(评价\)', '', col).strip()] = s2_row[col]
                        
                        scores.append(score_obj); observations.append(obs_obj)
                        
                    db_cache[month_key]["records"][rec_id] = {
                        "date": base_date_str, "mode": mode, "person": person,
                        "airport": ap_code, "is_shihang": is_shihang,
                        "daily_stats": daily_stats, "scores": scores, "observations": observations
                    }
                
        # 批量写回底层 JSON 数据库
        for m_key, db_data in db_cache.items():
            db_file = os.path.join(backup_path, f"backup_{m_key}.json")
            if os.path.exists(db_file):
                try:
                    with open(db_file, 'r', encoding='utf-8') as f:
                        existing = json.load(f)
                        existing['records'].update(db_data['records'])
                        db_data = existing
                except: pass
            with open(db_file, 'w', encoding='utf-8') as f:
                json.dump(db_data, f, ensure_ascii=False, indent=2)
                
        return jsonify({"success": True, "message": f"✅ 同步成功！从 Excel 云端成功提取并更新了 {len(db_cache)} 个月份的本地 JSON 数据库。"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/export_stats', methods=['POST'])
def export_stats_api():
    try:
        data = request.json
        excel_root = data.get('excel_root')
        table_data = data.get('table_data')
        title = data.get('title')
        time_type = data.get('time_type', 'day')
        base_date = data.get('base_date', '2026-01-01')
        
        if not excel_root: return jsonify({"success": False, "error": "请先配置 Excel 导出目录！"})
        
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        
        # 🌟 1. 智能分配文件夹 (年份 \ 月份)
        try:
            # 兼容多日期逗号分隔，取第一个日期作为归属
            base_dt = datetime.strptime(base_date.split(',')[0].strip(), '%Y-%m-%d')
        except:
            base_dt = datetime.now()
            
        y_str = f"{base_dt.year}年"
        m_str = f"{base_dt.month}月"
        
        # 年统计放年份里，月统计/日统计放月份里
        if time_type == 'year':
            folder = os.path.join(excel_root, y_str)
        else:
            folder = os.path.join(excel_root, y_str, m_str)
            
        os.makedirs(folder, exist_ok=True)
        file_path = os.path.join(folder, f"{title}.xlsx")
        
        # 🌟 2. 完美复刻前端双行复杂表格
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "统计汇总"
        
        font_bold = Font(name='仿宋_GB2312', size=12, bold=True)
        font_normal = Font(name='仿宋_GB2312', size=11)
        align_center = Alignment(horizontal='center', vertical='center')
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        first_col_name = "机场" if "机场预报" in title else "评定对象"
        headers = [first_col_name, "", "完美", "优秀", "空报", "漏报", "准确", "参评", "总评", "总分"]
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(1, i)
            cell.font = font_bold; cell.alignment = align_center; cell.border = border_thin

        row_idx = 2
        merges = []
        
        for row in table_data:
            obj_name = row.get("统计对象", "")
            
            # 计算复杂的总分逻辑
            acc_rate = (row.get("准确项次",0)/row.get("总评项次",1)) * 100 if row.get("总评项次",0)>0 else 0
            perf_rate = (row.get("完美项次",0)/row.get("参评项次",1)) * 100 if row.get("参评项次",0)>0 else 0
            exc_rate = (row.get("优秀项次",0)/row.get("参评项次",1)) * 100 if row.get("参评项次",0)>0 else 0
            fa_rate = (row.get("空报项次",0)/row.get("总评项次",1)) * 100 if row.get("总评项次",0)>0 else 0
            miss_rate = (row.get("漏报项次",0)/row.get("总评项次",1)) * 100 if row.get("总评项次",0)>0 else 0
            score = (acc_rate*0.5) + (perf_rate*0.3) + (exc_rate*0.2) - (fa_rate*0.1) - (miss_rate*0.1)
            
            row1 = [obj_name, "次数", row.get("完美项次",0), row.get("优秀项次",0), row.get("空报项次",0), row.get("漏报项次",0), row.get("准确项次",0), row.get("参评项次",0), row.get("总评项次",0), f"{score:.2f}"]
            row2 = ["", "概率", row.get("完美率","/"), row.get("优秀率","/"), row.get("空报率","/"), row.get("漏报率","/"), row.get("准确率","/"), "", "", ""]
            
            ws.append(row1)
            ws.append(row2)
            
            # 首列与尾列的上下合并
            merges.append((row_idx, 1, row_idx+1, 1))
            merges.append((row_idx, 10, row_idx+1, 10))
            
            for r in range(row_idx, row_idx+2):
                for c in range(1, 11):
                    cell = ws.cell(r, c)
                    cell.font = font_normal; cell.alignment = align_center; cell.border = border_thin
            row_idx += 2
            
        # 表尾规则说明合并
        desc_row = ["评定逻辑：总分=(准确率*50% + 完美率*30% + 优秀率*20% - 空报率*10% - 漏报率*10%)"] + [""]*9
        ws.append(desc_row)
        merges.append((row_idx, 1, row_idx, 10))
        desc_cell = ws.cell(row_idx, 1)
        desc_cell.font = font_bold; desc_cell.alignment = Alignment(horizontal='left', vertical='center'); desc_cell.border = border_thin
        for c in range(2, 11): ws.cell(row_idx, c).border = border_thin
            
        # 统一执行合并单元格
        for m in merges:
            ws.merge_cells(start_row=m[0], start_column=m[1], end_row=m[2], end_column=m[3])
            
        # 优化列宽
        ws.column_dimensions['A'].width = 15
        for col_letter in ['B','C','D','E','F','G','H','I','J']: ws.column_dimensions[col_letter].width = 12
        
        wb.save(file_path)
        return jsonify({"success": True, "message": f"成功导出至 {file_path}"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500