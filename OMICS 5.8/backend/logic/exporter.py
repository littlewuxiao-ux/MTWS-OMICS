import os
import json
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import re

# 映射评分项到标准名称
ITEM_MAP = {
    '最大风速(MPS)': '大风类', '最差能见度(m)': '低能见度', '最低云高(m)': '低云', 
    '雷雨类': '雷雨类', '强降水(无雷)类': '强降雨类(无雷)', '积冰类': '积冰类', '特殊类': '特殊类'
}

def process_stats_and_save(results, currentMode, backup_path, excel_root, eval_person, base_date_str, rater="未知"):
    # 🌟 修复核心 1：彻底清洗人员名字、路径中可能携带的隐藏空格和换行符
    eval_person = re.sub(r'[\r\n\t\s]', '', str(eval_person)) if eval_person else "未知人员"
    excel_root = os.path.normpath(str(excel_root).strip())
    
    # =========================================================
    # 🌟 需求4：彻底以“预报日期”为准！从报文时次逆向推算真实日期
    actual_date_str = base_date_str
    try:
        for k, d in results.items():
            if d.get('scores'):
                first_time = d['scores'][0].get('时次', '')
                if len(first_time) >= 2:
                    fcst_day = int(first_time[:2])
                    base_dt = datetime.strptime(base_date_str, '%Y-%m-%d')
                    fcst_m = base_dt.month
                    fcst_y = base_dt.year
                    
                    # 跨月精准判断：如果报文日期跟UI日期def process_stats_and_save(results, curr差了15天以上，肯定是跨月了
                    if fcst_day > base_dt.day + 15:
                        fcst_m -= 1
                        if fcst_m == 0: fcst_m = 12; fcst_y -= 1
                    elif fcst_day < base_dt.day - 15:
                        fcst_m += 1
                        if fcst_m == 13: fcst_m = 1; fcst_y += 1
                        
                    actual_date_str = f"{fcst_y}-{fcst_m:02d}-{fcst_day:02d}"
                break
    except Exception: pass
    
    base_date_str = actual_date_str # 强行覆盖外部 UI 传入的日期！
    # =========================================================

    try: dt = datetime.strptime(base_date_str, '%Y-%m-%d')
    except: dt = datetime.now()
    mmdd = dt.strftime('%m%d')

    os.makedirs(backup_path, exist_ok=True)
    db_file = os.path.join(backup_path, f"backup_{dt.strftime('%Y%m')}.json")
    if os.path.exists(db_file):
        try:
            with open(db_file, 'r', encoding='utf-8') as f: db_data = json.load(f)
        except: db_data = {"month": dt.strftime('%Y%m'), "records": {}}
    else: db_data = {"month": dt.strftime('%Y%m'), "records": {}}

    first_airport_code = "UNKNOWN"
    airport_all_scores = {}

    for key, data in results.items():
        ap_code = key.split('_')[0]
        if first_airport_code == "UNKNOWN": first_airport_code = ap_code
        scores = data.get('scores', [])
        if not scores: continue
        
        if ap_code not in airport_all_scores: airport_all_scores[ap_code] = []
        airport_all_scores[ap_code].extend(scores)

        total_slots = len(scores) 
        cat_stats = {}
        any_weather_evaluated = False
        ordered_cats = ['大风类', '低能见度', '低云', '雷雨类', '强降雨类(无雷)', '积冰类', '特殊类']
        
        for cat in ordered_cats: cat_stats[cat] = {"完美": 0, "优秀": 0, "空报": 0, "漏报": 0, "准确": 0, "参评": 0, "总评": 0}

        items = [k for k in scores[0].keys() if k not in ('时次', '预报内容', '预报时效', '预报全文')]
        for raw_item in items:
            cat_name = ITEM_MAP.get(raw_item, raw_item)
            if cat_name not in ordered_cats: continue
            vals = [row.get(raw_item, '不评') for row in scores]
            perf, exc, fa, miss = vals.count('完美'), vals.count('优秀'), vals.count('空报'), vals.count('漏报')
            not_eval = vals.count('不评')
            eval_c = perf + exc + fa + miss 
            acc_c = perf + exc + not_eval        
            
            if eval_c > 0: 
                any_weather_evaluated = True
                cat_stats[cat_name] = {"完美": perf, "优秀": exc, "空报": fa, "漏报": miss, "准确": acc_c, "参评": eval_c, "总评": total_slots}

        rec_id = f"{base_date_str}_{currentMode}_{eval_person}_{key}"
        db_data["records"][rec_id] = {
            "date": base_date_str, "mode": currentMode, "person": eval_person, 
            "airport": ap_code, "is_shihang": not any_weather_evaluated, 
            "daily_stats": cat_stats, "scores": scores, "observations": data.get('observations', [])
        }

    # === Excel 导出部分 ===
    sheet1_data = []
    merges = []
    row_offset = 1

    if currentMode == 'manual':
        sheet1_data.append([f"评分人：{rater}    评定对象：{eval_person}"] + [""] * 9)
        merges.append((row_offset, 1, row_offset, 10))
        row_offset += 1

    sheet1_data.append(['机场', '天气分类', '统计', '完美', '优秀', '空报', '漏报', '准确', '参评', '总评'])
    row_offset += 1

    for ap_code, all_scores in airport_all_scores.items():
        start_ap_row = row_offset
        total_slots = len(all_scores)
        
        cat_stats = {cat: {"完美":0, "优秀":0, "空报":0, "漏报":0, "准确":0, "参评":0, "总评":0} for cat in ordered_cats}
        any_evaluated = False
        
        items = [k for k in data.get('scores', [])[0].keys() if k not in ('时次', '预报内容', '预报时效', '预报全文')] if data.get('scores') else []
        for raw_item in items:
            cat_name = ITEM_MAP.get(raw_item, raw_item)
            if cat_name not in ordered_cats: continue
            vals = [r.get(raw_item, '不评') for r in all_scores]
            perf, exc, fa, miss = vals.count('完美'), vals.count('优秀'), vals.count('空报'), vals.count('漏报')
            not_eval = vals.count('不评')
            eval_c = perf + exc + fa + miss
            acc_c = perf + exc + not_eval
            
            if eval_c > 0: 
                any_evaluated = True
                cat_stats[cat_name]["完美"] += perf; cat_stats[cat_name]["优秀"] += exc
                cat_stats[cat_name]["空报"] += fa; cat_stats[cat_name]["漏报"] += miss
                cat_stats[cat_name]["准确"] += acc_c; cat_stats[cat_name]["参评"] += eval_c
                cat_stats[cat_name]["总评"] += total_slots

        sum_perf = sum_exc = sum_fa = sum_miss = sum_acc = sum_eval = sum_total = 0

        for cat in ordered_cats:
            s = cat_stats[cat]
            ev = s["参评"]; tot = s["总评"]
            
            row1 = [ap_code, cat, '次数', s['完美'], s['优秀'], s['空报'], s['漏报'], s['准确'], ev, tot]
            row2 = ["", "", '概率', 
                    f"{(s['完美']/ev*100):.2f}%" if ev>0 else "/", f"{(s['优秀']/ev*100):.2f}%" if ev>0 else "/",
                    f"{(s['空报']/tot*100):.2f}%" if tot>0 else "/", f"{(s['漏报']/tot*100):.2f}%" if tot>0 else "/",
                    f"{(s['准确']/tot*100):.2f}%" if tot>0 else "/", "", ""]
            
            sheet1_data.append(row1); sheet1_data.append(row2)
            merges.append((row_offset, 2, row_offset+1, 2))
            
            sum_perf += s['完美']; sum_exc += s['优秀']; sum_fa += s['空报']; sum_miss += s['漏报']
            sum_acc += s['准确']; sum_eval += ev; sum_total += tot
            row_offset += 2
        
        if not any_evaluated:
            row1 = [ap_code, '适航', '次数', 0, 0, 0, 0, total_slots, 0, total_slots]
            f_zero = "0.00%" if total_slots>0 else "/"
            row2 = ["", "", '概率', "/", "/", f_zero, f_zero, f"{(total_slots/total_slots*100):.2f}%" if total_slots>0 else "/", "", ""]
        else:
            row1 = [ap_code, '总计', '次数', sum_perf, sum_exc, sum_fa, sum_miss, sum_acc, sum_eval, sum_total]
            row2 = ["", "", '概率',
                    f"{(sum_perf/sum_eval*100):.2f}%" if sum_eval>0 else "/", f"{(sum_exc/sum_eval*100):.2f}%" if sum_eval>0 else "/",
                    f"{(sum_fa/sum_total*100):.2f}%" if sum_total>0 else "/", f"{(sum_miss/sum_total*100):.2f}%" if sum_total>0 else "/",
                    f"{(sum_acc/sum_total*100):.2f}%" if sum_total>0 else "/", "", ""]
            
        sheet1_data.append(row1); sheet1_data.append(row2)
        merges.append((row_offset, 2, row_offset+1, 2))
        row_offset += 2
        merges.append((start_ap_row, 1, row_offset-1, 1))

    # === Sheet2 数据明细 ===
    sheet2_all_details = []
    ordered_items = ['最大风速(MPS)', '最差能见度(m)', '最低云高(m)', '雷雨类', '强降水(无雷)类', '积冰类', '特殊类']
    
    for key, data in results.items():
        ap_code = key.split('_')[0]
        obs_dict = {row['时次']: row for row in data.get('observations', [])}
        
        # 🌟 修复需求1：强制锁定 Excel 明细表提取的顺序
        if data.get('scores'):
            available_keys = data.get('scores')[0].keys()
            items = [k for k in ordered_items if k in available_keys]
        else:
            items = []
            
        for score_row in data.get('scores', []):
            t = score_row['时次']
            obs_row = obs_dict.get(t, {})
            row_cp = {
                "机场": ap_code, 
                "预报时效": score_row.get("预报时效", "/"),
                "时次": t, 
                "预报全文": score_row.get("预报全文", "/")
            }
            for item_key in items:
                # 🌟 双重锁死：绝对不允许这几个词进入 实况/评分 的拆分循环
                if item_key in ('时次', '预报内容', '预报时效', '预报全文', '_time'): continue
                row_cp[f"{item_key}(实况)"] = obs_row.get(item_key, '/')
                row_cp[f"{item_key}(评分)"] = score_row[item_key]
            sheet2_all_details.append(row_cp)

    if currentMode == 'manual':
        folder = os.path.join(excel_root, f"{dt.year}年", f"{dt.month}月", eval_person)
        filename = f"席位预报质量评价表-{eval_person}-{mmdd}.xlsx"
    else:
        folder = os.path.join(excel_root, f"{dt.year}年", f"{dt.month}月")
        filename = f"机场预报质量评定-{mmdd}.xlsx"

    # 🌟 修复核心 3：再次整体规范化目标文件夹路径
    folder = os.path.normpath(folder)
    full_path = os.path.join(folder, filename)

    try:
        # 🌟 修复核心 4：将建文件夹的操作移入 try 块！
        # 如果 Z 盘此刻掉线、权限不足，或者路径被占用，就会被抛出具体的错误原因，不再是含糊的"Errno 2"
        os.makedirs(folder, exist_ok=True)
        
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "日统计"

        font_fangsong = Font(name='仿宋_GB2312', size=11)
        font_bold = Font(name='仿宋_GB2312', size=12, bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 18
        for col_letter in ['C','D','E','F','G','H','I','J']: ws1.column_dimensions[col_letter].width = 12

        for r_idx, row in enumerate(sheet1_data, 1):
            for c_idx, val in enumerate(row, 1):
                c = ws1.cell(r_idx, c_idx, val)
                c.font = font_bold if r_idx <= (2 if currentMode=='manual' else 1) else font_fangsong
                c.alignment = align_center
                c.border = border_thin

        for m in merges: ws1.merge_cells(start_row=m[0], start_column=m[1], end_row=m[2], end_column=m[3])

        ws2 = wb.create_sheet(title="要素分解与汇总")
        if sheet2_all_details:
            headers2 = list(sheet2_all_details[0].keys())
            ws2.append(headers2)
            for row_dict in sheet2_all_details: ws2.append([row_dict.get(h, '') for h in headers2])
            for r in ws2.iter_rows():
                for cell in r:
                    cell.font = font_fangsong
                    cell.alignment = align_center
                    cell.border = border_thin

        wb.save(full_path)
    except Exception as e:
        raise Exception(f"路径 [{full_path}] 无法写入。\n可能原因：Z盘未连接、权限不足或文件名有误。\n系统报错: {str(e)}")
    
    with open(db_file, 'w', encoding='utf-8') as f: json.dump(db_data, f, ensure_ascii=False, indent=2)